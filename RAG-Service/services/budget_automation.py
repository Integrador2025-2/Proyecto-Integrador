import os
import json
from typing import List, Dict, Any, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import logging

from .rag_service import RAGService
from .llm_service import LLMService
from .budget_extractor import BudgetExtractor

logger = logging.getLogger(__name__)

class BudgetAutomationService:
    """Servicio para automatización de presupuestos basado en RAG"""
    
    def __init__(self):
        self.rag_service = RAGService()
        self.budget_extractor = BudgetExtractor()
        
        # Inicializar servicio LLM (opcional)
        try:
            self.llm_service = LLMService()
            self.use_llm = True
        except Exception as e:
            self.llm_service = None
            self.use_llm = False
            logger.warning(f"LLM no disponible para generación de presupuestos. Error: {str(e)}")
        
        # Categorías de presupuesto mapeadas a los rubros del sistema
        self.budget_categories = {
            "TalentoHumano": {
                "description": "Recursos humanos, salarios, honorarios",
                "keywords": ["personal", "salario", "honorario", "recurso humano", "trabajador", "empleado"]
            },
            "ServiciosTecnologicos": {
                "description": "Servicios de tecnología, consultoría técnica",
                "keywords": ["servicio", "tecnología", "consultoría", "desarrollo", "software", "sistema"]
            },
            "EquiposSoftware": {
                "description": "Equipos de cómputo, software, licencias",
                "keywords": ["equipo", "computadora", "software", "licencia", "hardware", "tecnología"]
            },
            "MaterialesInsumos": {
                "description": "Materiales, insumos, suministros",
                "keywords": ["material", "insumo", "suministro", "herramienta", "equipo", "consumible"]
            },
            "CapacitacionEventos": {
                "description": "Capacitaciones, eventos, talleres",
                "keywords": ["capacitación", "evento", "taller", "curso", "entrenamiento", "formación"]
            },
            "GastosViaje": {
                "description": "Gastos de viaje, transporte, hospedaje",
                "keywords": ["viaje", "transporte", "hospedaje", "desplazamiento", "movilización", "viático"]
            }
        }
    
    async def generate_budget(self, project_id: int, project_description: str, 
                           budget_categories: List[str], duration_years: int,
                           activities: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
        """
        Generar presupuesto automáticamente basado en documentos del proyecto y/o actividades
        
        Args:
            project_id: ID del proyecto
            project_description: Descripción del proyecto
            budget_categories: Lista de categorías de presupuesto
            duration_years: Duración del proyecto en años
            activities: Lista opcional de actividades del proyecto
        """
        try:
            # PASO 1: Intentar extraer presupuesto inteligentemente desde documentos Excel/DOCX del proyecto
            logger.info(f"Iniciando generación de presupuesto para proyecto {project_id}")
            
            extracted_budget = await self._extract_budget_from_project_documents(project_id)
            
            # Si se extrajo presupuesto completo desde documentos, usarlo
            if extracted_budget and extracted_budget.get("total_activities", 0) > 0:
                logger.info(f"Presupuesto extraído exitosamente: {extracted_budget['total_activities']} actividades")
                
                # Verificar si necesita completar valores faltantes con LLM
                summary = self.budget_extractor.calculate_budget_summary(extracted_budget["activities"])
                
                if summary["needs_llm_generation"] and self.use_llm and self.llm_service:
                    logger.info(f"Completando {summary['activities_without_budget']} actividades sin presupuesto usando LLM")
                    extracted_budget = await self._complete_missing_budgets_with_llm(
                        extracted_budget, project_id, project_description, duration_years
                    )
                
                # Convertir a formato final y generar Excel
                budget_data = await self._convert_extracted_to_budget_data(
                    extracted_budget, project_id, project_description, duration_years
                )
                
                excel_path = await self._generate_excel_budget(budget_data, project_id)
                
                return {
                    "project_id": project_id,
                    "budget_data": budget_data,
                    "excel_path": excel_path,
                    "generated_at": datetime.now().isoformat(),
                    "confidence_score": extracted_budget.get("confidence", 0.85),
                    "source_documents": len(extracted_budget.get("source_files", [])),
                    "source_activities": extracted_budget["total_activities"],
                    "method": "extracted_from_documents"
                }
            
            # PASO 2: Si hay actividades proporcionadas y LLM disponible, usar generación inteligente con LLM
            if activities and self.use_llm and self.llm_service:
                logger.info(f"Generando presupuesto con LLM basado en {len(activities)} actividades")
                return await self._generate_budget_with_llm(
                    project_id, project_description, activities, budget_categories, duration_years
                )
            
            # PASO 3: Método tradicional basado en análisis RAG de documentos
            logger.info("Usando método tradicional basado en RAG")
            project_docs = await self.rag_service.get_project_documents(project_id)
            
            if not project_docs:
                logger.warning("No se encontraron documentos, generando presupuesto por defecto")
                return await self._generate_default_budget(project_description, budget_categories, duration_years)
            
            # Extraer información de presupuesto de los documentos
            budget_info = await self._extract_budget_information(project_docs, budget_categories)
            
            # Generar presupuesto detallado
            budget_data = await self._create_detailed_budget(
                project_id, project_description, budget_info, budget_categories, duration_years
            )
            
            # Generar archivo Excel
            excel_path = await self._generate_excel_budget(budget_data, project_id)
            
            return {
                "project_id": project_id,
                "budget_data": budget_data,
                "excel_path": excel_path,
                "generated_at": datetime.now().isoformat(),
                "confidence_score": budget_data.get("confidence_score", 0.0),
                "source_documents": [doc.get("filename", "unknown") for doc in project_docs] if project_docs else [],
                "source_activities": []
            }
            
        except Exception as e:
            logger.error(f"Error generando presupuesto: {str(e)}")
            raise Exception(f"Error generando presupuesto: {str(e)}")
    
    async def _generate_budget_with_llm(
        self,
        project_id: int,
        project_description: str,
        activities: List[Dict[str, Any]],
        budget_categories: List[str],
        duration_years: int
    ) -> Dict[str, Any]:
        """Generar presupuesto usando LLM basado en actividades"""
        try:
            # Obtener contexto de documentos del proyecto si están disponibles
            project_docs_context = None
            project_docs_count = 0
            try:
                project_docs = await self.rag_service.get_project_documents(project_id)
                if project_docs:
                    project_docs_count = len(project_docs)
                    # Combinar contenido de documentos para contexto
                    project_docs_context = " ".join([
                        " ".join([chunk["content"] for chunk in doc["chunks"]])
                        for doc in project_docs[:5]  # Limitar a 5 documentos
                    ])
            except:
                pass
            
            # Generar presupuesto con LLM
            budget_data = await self.llm_service.generate_budget_from_activities(
                project_description=project_description,
                activities=activities,
                project_documents_context=project_docs_context,
                duration_years=duration_years,
                budget_categories=budget_categories
            )
            
            # Agregar información adicional
            budget_data["project_id"] = project_id
            budget_data["project_description"] = project_description
            budget_data["duration_years"] = duration_years
            
            # Generar archivo Excel
            excel_path = await self._generate_excel_budget(budget_data, project_id)
            
            return {
                "project_id": project_id,
                "budget_data": budget_data,
                "excel_path": excel_path,
                "generated_at": budget_data.get("generated_at", datetime.now().isoformat()),
                "confidence_score": budget_data.get("confidence_score", 0.8),
                "source_documents": project_docs_count,
                "source_activities": len(activities),
                "method": "llm_based"
            }
            
        except Exception as e:
            raise Exception(f"Error generando presupuesto con LLM: {str(e)}")
    
    async def get_budget_suggestions(self, project_id: int, category: str = None) -> List[Dict[str, Any]]:
        """Obtener sugerencias de presupuesto para un proyecto específico"""
        try:
            # Obtener documentos del proyecto
            project_docs = await self.rag_service.get_project_documents(project_id)
            
            if not project_docs:
                return []
            
            suggestions = []
            categories_to_analyze = [category] if category else list(self.budget_categories.keys())
            
            for cat in categories_to_analyze:
                if cat in self.budget_categories:
                    suggestion = await self._generate_category_suggestion(project_docs, cat)
                    if suggestion:
                        suggestions.append(suggestion)
            
            return suggestions
            
        except Exception as e:
            raise Exception(f"Error obteniendo sugerencias: {str(e)}")
    
    async def _extract_budget_information(self, project_docs: List[Dict[str, Any]], 
                                        budget_categories: List[str]) -> Dict[str, Any]:
        """Extraer información de presupuesto de los documentos del proyecto"""
        budget_info = {}
        
        for doc in project_docs:
            content = " ".join([chunk["content"] for chunk in doc["chunks"]])
            
            for category in budget_categories:
                if category not in budget_info:
                    budget_info[category] = {
                        "items": [],
                        "total_estimated": 0.0,
                        "confidence": 0.0
                    }
                
                # Buscar información específica de la categoría
                category_info = await self._analyze_category_content(content, category)
                if category_info:
                    budget_info[category]["items"].extend(category_info["items"])
                    budget_info[category]["total_estimated"] += category_info["total_estimated"]
                    budget_info[category]["confidence"] = max(
                        budget_info[category]["confidence"], 
                        category_info["confidence"]
                    )
        
        return budget_info
    
    async def _analyze_category_content(self, content: str, category: str) -> Optional[Dict[str, Any]]:
        """Analizar contenido para extraer información de una categoría específica"""
        category_config = self.budget_categories.get(category)
        if not category_config:
            return None
        
        # Buscar menciones de la categoría en el contenido
        content_lower = content.lower()
        keywords = category_config["keywords"]
        
        # Calcular relevancia basada en palabras clave
        relevance_score = sum(1 for keyword in keywords if keyword in content_lower) / len(keywords)
        
        if relevance_score < 0.1:  # Umbral mínimo de relevancia
            return None
        
        # Extraer información de costos mencionados
        cost_items = self._extract_cost_items(content, category)
        
        return {
            "items": cost_items,
            "total_estimated": sum(item.get("estimated_cost", 0) for item in cost_items),
            "confidence": relevance_score
        }
    
    def _extract_cost_items(self, content: str, category: str) -> List[Dict[str, Any]]:
        """Extraer elementos de costo del contenido"""
        items = []
        
        # Buscar patrones de costos (números seguidos de palabras como "pesos", "dólares", etc.)
        import re
        
        # Patrón para encontrar costos
        cost_patterns = [
            r'(\d+(?:,\d{3})*(?:\.\d{2})?)\s*(?:pesos?|dólares?|usd|cop|colombianos?)',
            r'costo[:\s]+(\d+(?:,\d{3})*(?:\.\d{2})?)',
            r'precio[:\s]+(\d+(?:,\d{3})*(?:\.\d{2})?)',
            r'valor[:\s]+(\d+(?:,\d{3})*(?:\.\d{2})?)'
        ]
        
        for pattern in cost_patterns:
            matches = re.finditer(pattern, content, re.IGNORECASE)
            for match in matches:
                try:
                    cost_str = match.group(1).replace(',', '')
                    cost = float(cost_str)
                    
                    # Obtener contexto alrededor del costo
                    start = max(0, match.start() - 50)
                    end = min(len(content), match.end() + 50)
                    context = content[start:end].strip()
                    
                    items.append({
                        "description": context,
                        "estimated_cost": cost,
                        "category": category,
                        "confidence": 0.7
                    })
                except ValueError:
                    continue
        
        return items
    
    async def _create_detailed_budget(self, project_id: int, project_description: str,
                                    budget_info: Dict[str, Any], budget_categories: List[str],
                                    duration_years: int) -> Dict[str, Any]:
        """Crear presupuesto detallado basado en la información extraída"""
        
        budget_data = {
            "project_id": project_id,
            "project_description": project_description,
            "duration_years": duration_years,
            "categories": {},
            "total_budget": 0.0,
            "confidence_score": 0.0,
            "generated_at": datetime.now().isoformat()
        }
        
        total_confidence = 0.0
        category_count = 0
        
        for category in budget_categories:
            if category in budget_info and budget_info[category]["items"]:
                category_data = budget_info[category]
                
                # Generar items detallados para la categoría
                detailed_items = await self._generate_detailed_category_items(
                    category, category_data, duration_years
                )
                
                budget_data["categories"][category] = {
                    "description": self.budget_categories[category]["description"],
                    "items": detailed_items,
                    "subtotal": sum(item["total_cost"] for item in detailed_items),
                    "confidence": category_data["confidence"]
                }
                
                budget_data["total_budget"] += budget_data["categories"][category]["subtotal"]
                total_confidence += category_data["confidence"]
                category_count += 1
        
        # Calcular confianza promedio
        if category_count > 0:
            budget_data["confidence_score"] = total_confidence / category_count
        
        return budget_data
    
    async def _generate_detailed_category_items(self, category: str, category_data: Dict[str, Any],
                                             duration_years: int) -> List[Dict[str, Any]]:
        """Generar items detallados para una categoría específica"""
        items = []
        
        # Si hay items extraídos de los documentos, usarlos como base
        if category_data["items"]:
            for item in category_data["items"]:
                items.append({
                    "description": item["description"],
                    "quantity": 1,
                    "unit_cost": item["estimated_cost"],
                    "total_cost": item["estimated_cost"],
                    "justification": f"Basado en análisis de documentos del proyecto",
                    "year": 1,
                    "confidence": item.get("confidence", 0.7)
                })
        else:
            # Generar items por defecto basados en la categoría
            default_items = self._get_default_category_items(category, duration_years)
            items.extend(default_items)
        
        return items
    
    def _get_default_category_items(self, category: str, duration_years: int) -> List[Dict[str, Any]]:
        """Obtener items por defecto para una categoría cuando no hay información específica"""
        default_items = {
            "TalentoHumano": [
                {
                    "description": "Coordinador del proyecto",
                    "quantity": 1,
                    "unit_cost": 5000000,
                    "total_cost": 5000000 * duration_years,
                    "justification": "Recurso humano principal para la coordinación del proyecto",
                    "year": 1,
                    "confidence": 0.5
                }
            ],
            "ServiciosTecnologicos": [
                {
                    "description": "Servicios de desarrollo de software",
                    "quantity": 1,
                    "unit_cost": 2000000,
                    "total_cost": 2000000 * duration_years,
                    "justification": "Servicios tecnológicos necesarios para el proyecto",
                    "year": 1,
                    "confidence": 0.5
                }
            ],
            "EquiposSoftware": [
                {
                    "description": "Equipos de cómputo",
                    "quantity": 2,
                    "unit_cost": 3000000,
                    "total_cost": 6000000,
                    "justification": "Equipos necesarios para el desarrollo del proyecto",
                    "year": 1,
                    "confidence": 0.5
                }
            ],
            "MaterialesInsumos": [
                {
                    "description": "Materiales y suministros generales",
                    "quantity": 1,
                    "unit_cost": 1000000,
                    "total_cost": 1000000 * duration_years,
                    "justification": "Materiales básicos para el proyecto",
                    "year": 1,
                    "confidence": 0.5
                }
            ],
            "CapacitacionEventos": [
                {
                    "description": "Capacitación del equipo",
                    "quantity": 1,
                    "unit_cost": 1500000,
                    "total_cost": 1500000,
                    "justification": "Capacitación necesaria para el equipo del proyecto",
                    "year": 1,
                    "confidence": 0.5
                }
            ],
            "GastosViaje": [
                {
                    "description": "Gastos de viaje y desplazamiento",
                    "quantity": 1,
                    "unit_cost": 2000000,
                    "total_cost": 2000000 * duration_years,
                    "justification": "Gastos de desplazamiento para actividades del proyecto",
                    "year": 1,
                    "confidence": 0.5
                }
            ]
        }
        
        return default_items.get(category, [])
    
    async def _generate_excel_budget(self, budget_data: Dict[str, Any], project_id: int) -> str:
        """Generar archivo Excel con el presupuesto"""
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Presupuesto del Proyecto"
            
            # Configurar estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            currency_format = '"$"#,##0'
            
            # Encabezado
            ws['A1'] = f"Presupuesto del Proyecto - ID: {project_id}"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:F1')
            
            ws['A2'] = f"Descripción: {budget_data['project_description']}"
            ws.merge_cells('A2:F2')
            
            ws['A3'] = f"Duración: {budget_data['duration_years']} año(s)"
            ws.merge_cells('A3:F3')
            
            # Espacio
            current_row = 5
            
            # Por cada categoría
            for category, data in budget_data['categories'].items():
                # Título de la categoría
                ws[f'A{current_row}'] = f"{category} - {data['description']}"
                ws[f'A{current_row}'].font = Font(bold=True, size=12)
                ws.merge_cells(f'A{current_row}:F{current_row}')
                current_row += 1
                
                # Encabezados de la tabla
                headers = ['Descripción', 'Cantidad', 'Costo Unitario', 'Costo Total', 'Justificación', 'Año']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                current_row += 1
                
                # Items de la categoría
                for item in data['items']:
                    ws[f'A{current_row}'] = item['description']
                    ws[f'B{current_row}'] = item['quantity']
                    ws[f'C{current_row}'] = item['unit_cost']
                    ws[f'C{current_row}'].number_format = currency_format
                    ws[f'D{current_row}'] = item['total_cost']
                    ws[f'D{current_row}'].number_format = currency_format
                    ws[f'E{current_row}'] = item['justification']
                    ws[f'F{current_row}'] = item['year']
                    current_row += 1
                
                # Subtotal de la categoría
                ws[f'A{current_row}'] = f"Subtotal {category}:"
                ws[f'A{current_row}'].font = Font(bold=True)
                ws[f'D{current_row}'] = data['subtotal']
                ws[f'D{current_row}'].font = Font(bold=True)
                ws[f'D{current_row}'].number_format = currency_format
                current_row += 2
            
            # Total general
            ws[f'A{current_row}'] = "TOTAL GENERAL:"
            ws[f'A{current_row}'].font = Font(bold=True, size=14)
            ws[f'D{current_row}'] = budget_data['total_budget']
            ws[f'D{current_row}'].font = Font(bold=True, size=14)
            ws[f'D{current_row}'].number_format = currency_format
            
            # Ajustar ancho de columnas
            column_widths = [50, 10, 15, 15, 40, 8]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # Guardar archivo
            filename = f"presupuesto_proyecto_{project_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"./generated_budgets/{filename}"
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            wb.save(filepath)
            return filepath
            
        except Exception as e:
            raise Exception(f"Error generando archivo Excel: {str(e)}")
    
    async def _generate_default_budget(self, project_description: str, 
                                     budget_categories: List[str], duration_years: int) -> Dict[str, Any]:
        """Generar presupuesto por defecto cuando no hay documentos"""
        budget_data = {
            "project_description": project_description,
            "duration_years": duration_years,
            "categories": {},
            "total_budget": 0.0,
            "confidence_score": 0.3,  # Baja confianza para presupuesto por defecto
            "generated_at": datetime.now().isoformat()
        }
        
        for category in budget_categories:
            default_items = self._get_default_category_items(category, duration_years)
            subtotal = sum(item["total_cost"] for item in default_items)
            
            budget_data["categories"][category] = {
                "description": self.budget_categories[category]["description"],
                "items": default_items,
                "subtotal": subtotal,
                "confidence": 0.3
            }
            
            budget_data["total_budget"] += subtotal
        
        return budget_data
    
    async def _generate_category_suggestion(self, project_docs: List[Dict[str, Any]], 
                                          category: str) -> Optional[Dict[str, Any]]:
        """Generar sugerencia específica para una categoría"""
        try:
            # Buscar información relevante en los documentos
            relevant_content = []
            for doc in project_docs:
                content = " ".join([chunk["content"] for chunk in doc["chunks"]])
                if any(keyword in content.lower() for keyword in self.budget_categories[category]["keywords"]):
                    relevant_content.append(content)
            
            if not relevant_content:
                return None
            
            # Generar sugerencia basada en el contenido
            suggestion = {
                "category": category,
                "suggested_items": [],
                "reasoning": f"Basado en análisis de {len(relevant_content)} documentos del proyecto",
                "confidence": 0.6
            }
            
            # Extraer items de costo del contenido relevante
            for content in relevant_content:
                cost_items = self._extract_cost_items(content, category)
                suggestion["suggested_items"].extend(cost_items)
            
            return suggestion
            
        except Exception as e:
            return None
    
    async def _extract_budget_from_project_documents(self, project_id: int) -> Optional[Dict[str, Any]]:
        """
        Extraer presupuesto inteligentemente desde archivos Excel/DOCX del proyecto.
        Busca documentos con presupuestos cargados en ChromaDB.
        """
        try:
            # Obtener documentos del proyecto
            project_docs = await self.rag_service.get_project_documents(project_id)
            
            if not project_docs:
                return None
            
            all_extracted_activities = []
            source_files = []
            
            # Buscar documentos Excel o DOCX que contengan presupuestos
            for doc in project_docs:
                filename = doc.get("filename", "")
                file_extension = os.path.splitext(filename)[1].lower()
                
                # Solo procesar Excel y DOCX
                if file_extension not in ['.xlsx', '.xls', '.docx']:
                    continue
                
                logger.info(f"Intentando extraer presupuesto de: {filename}")
                
                # Aquí necesitaríamos acceso al archivo original
                # Por ahora, intentamos extraer desde el contenido chunkeado en ChromaDB
                try:
                    # Reconstruir contenido del documento
                    full_content = " ".join([chunk["content"] for chunk in doc.get("chunks", [])])
                    
                    # Si el contenido contiene datos tabulares (detectar por patrones)
                    if self._contains_tabular_data(full_content):
                        # Parsear el contenido como tabla
                        extracted = await self._parse_tabular_content(full_content, filename)
                        if extracted and extracted.get("activities"):
                            all_extracted_activities.extend(extracted["activities"])
                            source_files.append(filename)
                    
                except Exception as e:
                    logger.warning(f"Error extrayendo de {filename}: {str(e)}")
                    continue
            
            if not all_extracted_activities:
                return None
            
            # Agrupar por rubro
            grouped_by_rubro = self.budget_extractor._group_activities_by_rubro(all_extracted_activities)
            
            return {
                "activities": all_extracted_activities,
                "grouped_by_rubro": grouped_by_rubro,
                "total_activities": len(all_extracted_activities),
                "rubros_found": list(grouped_by_rubro.keys()),
                "source_files": source_files,
                "extraction_method": "intelligent_from_chroma",
                "confidence": 0.8
            }
            
        except Exception as e:
            logger.error(f"Error extrayendo presupuesto de documentos: {str(e)}")
            return None
    
    def _contains_tabular_data(self, content: str) -> bool:
        """Detectar si el contenido tiene datos tabulares"""
        # Buscar patrones que indiquen tablas
        indicators = [
            "Hoja:",  # Indicador de Excel procesado
            "actividad",
            "rubro",
            "valor",
            "total"
        ]
        
        content_lower = content.lower()
        matches = sum(1 for ind in indicators if ind.lower() in content_lower)
        
        return matches >= 2
    
    async def _parse_tabular_content(self, content: str, filename: str) -> Optional[Dict[str, Any]]:
        """Parsear contenido tabular desde texto"""
        try:
            activities = []
            
            # Dividir por líneas
            lines = content.split('\n')
            
            # Buscar líneas que contengan información de presupuesto
            for line in lines:
                # Intentar extraer actividad y valor
                activity = await self._extract_activity_from_line(line, filename)
                if activity:
                    activities.append(activity)
            
            if not activities:
                return None
            
            return {
                "activities": activities,
                "source_file": filename
            }
            
        except Exception as e:
            logger.warning(f"Error parseando contenido tabular: {str(e)}")
            return None
    
    async def _extract_activity_from_line(self, line: str, source: str) -> Optional[Dict[str, Any]]:
        """Extraer actividad desde una línea de texto"""
        import re
        
        # Buscar patrones numéricos (valores de presupuesto)
        number_pattern = r'(\d{1,3}(?:[,\.]\d{3})*(?:[,\.]\d{2})?)'
        numbers = re.findall(number_pattern, line)
        
        if not numbers:
            return None
        
        # Identificar rubro
        rubro = self.budget_extractor.identify_rubro_from_text(line)
        if not rubro:
            return None
        
        # Extraer el valor más alto como total
        values = []
        for num_str in numbers:
            try:
                cleaned = num_str.replace(',', '').replace('.', '')
                if len(cleaned) > 2:
                    values.append(float(cleaned))
            except:
                continue
        
        if not values:
            return None
        
        total = max(values)
        
        return {
            "nombre": line[:100].strip(),  # Primeros 100 caracteres como nombre
            "rubro": rubro,
            "total": total,
            "cantidad": 1,
            "valor_unitario": total,
            "justificacion": f"Extraído de {source}",
            "especificaciones_tecnicas": "",
            "periodo": 1,
            "source_sheet": source,
            "has_budget_values": True
        }
    
    async def _complete_missing_budgets_with_llm(
        self,
        extracted_budget: Dict[str, Any],
        project_id: int,
        project_description: str,
        duration_years: int
    ) -> Dict[str, Any]:
        """Completar presupuestos faltantes usando LLM"""
        try:
            activities_to_complete = [
                act for act in extracted_budget["activities"]
                if not act.get("has_budget_values") or not act.get("total") or act["total"] <= 0
            ]
            
            if not activities_to_complete:
                return extracted_budget
            
            logger.info(f"Completando {len(activities_to_complete)} actividades con LLM")
            
            # Obtener contexto del proyecto
            project_docs_context = None
            try:
                project_docs = await self.rag_service.get_project_documents(project_id)
                if project_docs:
                    project_docs_context = " ".join([
                        " ".join([chunk["content"] for chunk in doc["chunks"]])
                        for doc in project_docs[:3]
                    ])[:3000]  # Limitar contexto
            except:
                pass
            
            # Generar presupuestos para las actividades faltantes
            completed_activities = await self.llm_service.generate_budget_from_activities(
                project_description=project_description,
                activities=activities_to_complete,
                project_documents_context=project_docs_context,
                duration_years=duration_years,
                budget_categories=list(extracted_budget["rubros_found"])
            )
            
            # Actualizar las actividades originales con los valores generados
            completed_map = {act["nombre"]: act for act in activities_to_complete}
            
            for category, data in completed_activities.get("categories", {}).items():
                for item in data.get("items", []):
                    activity_name = item.get("description", "")
                    if activity_name in completed_map:
                        completed_map[activity_name]["total"] = item.get("total_cost", 0)
                        completed_map[activity_name]["valor_unitario"] = item.get("unit_cost", 0)
                        completed_map[activity_name]["cantidad"] = item.get("quantity", 1)
                        completed_map[activity_name]["has_budget_values"] = True
            
            logger.info("Presupuestos completados exitosamente con LLM")
            
            return extracted_budget
            
        except Exception as e:
            logger.error(f"Error completando presupuestos con LLM: {str(e)}")
            return extracted_budget
    
    async def _convert_extracted_to_budget_data(
        self,
        extracted_budget: Dict[str, Any],
        project_id: int,
        project_description: str,
        duration_years: int
    ) -> Dict[str, Any]:
        """Convertir presupuesto extraído al formato final"""
        budget_data = {
            "project_id": project_id,
            "project_description": project_description,
            "duration_years": duration_years,
            "categories": {},
            "total_budget": 0.0,
            "confidence_score": extracted_budget.get("confidence", 0.85),
            "generated_at": datetime.now().isoformat(),
            "extraction_method": extracted_budget.get("extraction_method", "intelligent")
        }
        
        # Procesar actividades agrupadas por rubro
        for rubro, activities in extracted_budget.get("grouped_by_rubro", {}).items():
            items = []
            subtotal = 0.0
            
            for activity in activities:
                total = activity.get("total", 0) or 0
                cantidad = activity.get("cantidad", 1) or 1
                valor_unitario = activity.get("valor_unitario") or (total / cantidad if cantidad > 0 else 0)
                
                item = {
                    "description": activity.get("nombre", "Sin descripción"),
                    "quantity": cantidad,
                    "unit_cost": valor_unitario,
                    "total_cost": total,
                    "justification": activity.get("justificacion", "Extraído de documentos del proyecto"),
                    "year": activity.get("periodo", 1),
                    "confidence": 0.9 if activity.get("has_budget_values") else 0.6
                }
                
                items.append(item)
                subtotal += total
            
            if rubro in self.budget_categories:
                description = self.budget_categories[rubro]["description"]
            else:
                description = f"Categoría {rubro}"
            
            budget_data["categories"][rubro] = {
                "description": description,
                "items": items,
                "subtotal": subtotal,
                "confidence": 0.85
            }
            
            budget_data["total_budget"] += subtotal
        
        return budget_data
